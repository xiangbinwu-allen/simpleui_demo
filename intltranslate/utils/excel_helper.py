import json
import os
import re
from os.path import isfile

from openpyxl import load_workbook, Workbook

from intltranslate.utils import json_util


def _convert_excel_2_dict(file_path):
    dic = {}
    work_book = load_workbook(file_path)
    sheetnames = work_book.sheetnames
    if len(sheetnames) < 1:
        return
    sheet = work_book[sheetnames[0]]
    data_tuple = tuple(sheet.rows)
    for row in data_tuple:
        i = 0
        cell_values = []
        for cell in row:
            cell_values.append(cell.value)
            i += 1
        dic[cell_values[0]] = cell_values[2]
    return dic

def _convert_dict_2_json(json_file_name, dic_list):
    # 将dicnew转换为对象
    json_object = get_json(dic_list)

    # 将数据写入文件
    json.dump(json_object,json_file_name)
    with open(json_file_name,'wb') as f:
        f.write(str(json_object))


def conver_dic_2_dics(dic_list, dic_trun):
    for dic in dic_list:
        print(dic)
        print(dic.find("."))
        if dic.find(".") != -1:
            dic_trun_child = {}
            key = dic[:dic.index(".")]
            print(key)
            dic_trun[key] = dic_trun_child
            print(dic[dic.index(".")])
            conver_dic_2_dics(dic[dic.index(".")], dic_trun_child)
            continue
        dic_trun[dic] = dic_list[dic]
    print(dic_trun)


def save_2_excel(file_path, tables):
    if not file_path.endswith(".xlsx"):
        raise ValueError("文件必须为.xlsx结尾的Excel文件")
    work_book = Workbook()
    is_first = True
    for table in tables:
        if is_first:
            ws = work_book.active
            is_first = False
        else:
            ws = work_book.create_sheet()
        # 添加表头
        table_heads = []
        for attr in table[0].__dict__:
            # 过滤"_"开头的属性
            if not attr.startswith("_"):
                table_heads.append(attr)
        ws.append(table_heads)

        # 添加数据
        for row in table:
            data = []
            for head in table_heads:
                data.append(getattr(row, head))
            ws.append(data)
    try:
        # 生成保存文件夹路径
        folder_index = max(file_path.rfind("\\"), file_path.rfind("/"))
        if folder_index != -1:
            folder_path = file_path[0:folder_index]
            if not os.path.exists(folder_path):
                os.mkdir(folder_path)
        work_book.save(file_path)
    except Exception:
        raise OSError("创建Excel失败")

def _convert_value(value):
    """
    将单元格中数据，区分基本类型
    类似"true"/"false"(不区分大小写)转换为bool值
    长得像数字的转换为float类型
    其他（空格、空行）转换为None
    :param value: 单元格的值
    :return: 转换后的类型
    """
    value_str = str(value).lower()
    if value_str == 'true':
        return True
    elif value_str == 'false':
        return False
    elif re.match(r"^[+|-]?\d+.?\d*$", value_str):
        return float(value_str)
    elif re.match(r"^\s*$", value_str):
        return None
    else:
        return value

def get_json(i18nData):
    key_list = []
    value_list = []
    for i in i18nData:
        temp = i.split(".")
        key_list.append(temp)
        value_list.append(i18nData.get(i))
    json_data = {}
    for keys in key_list:
        generate_json(keys, value_list, json_data, 0)
    return json_data

def generate_json(keys, values, dict_data, index):
    # 判断是否已经在字典中有该字段
    if str(keys[index]) in dict_data:
        # 存在就在已存在的字典内的字典操作
        my_dict = dict_data.get(str(keys[index]))
        generate_json(keys, values, my_dict, index + 1)
    else:
        # 判断是否为最后一个字段
        if index + 1 == len(keys):
            # 如果是后一个字段就取值
            my_dict = {str(keys[index]): values[0]}
            values.pop(0)
        else:
            # 因为不是最后一个字段所以赋值一个空字典，防止出现None错误
            my_dict = {str(keys[index]): {}}
            generate_json(keys, values, my_dict, index)
        # temp = dict_data.get(str(keys[i]))
        dict_data.update(my_dict)


if __name__ == '__main__':
    file_path = os.path.join("C://Users//86136//Desktop//打包//EN-20210726-睿易app拓扑.xlsx")
    file_path2 = os.path.join(os.getcwd(), "templates\\app_file\\zh.json")
    dic={}
    _convert_dict_2_json(file_path2, dic)