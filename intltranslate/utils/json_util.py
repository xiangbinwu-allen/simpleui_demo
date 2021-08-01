import os
import json
import string
from pandas.core.frame import DataFrame
import os
import math
from openpyxl import load_workbook

from intltranslate.utils import excel_helper


def _update_file_data_from_excel(json_file_name, obj):
    # 1.读入**.json文件, 转为字典
    json_file_name = os.path.join(os.getcwd(), "templates\\app_file\\" + json_file_name)
    with open(json_file_name, 'r', encoding='UTF-8') as f:
        aa = json.load(f)
    dic=aa

    # 2.遍历字典, 转换字典为,父类1.子类2.key形式, 消除层级
    parent = ""
    split = "."
    dic_new={}
    turn_object_2_dict(dic, dic_new, parent, split)

    # 3.解析excel文件, 转为dict
    excel_dic = excel_helper._convert_excel_2_dict(obj)

    # 4.更新excel_dic数据到dicnew
    for key in excel_dic:
        dic_new[key]=excel_dic[key]

    # 5. 将dicnew转换为JSON, 写入文件 todo
    excel_helper._convert_dict_2_json(json_file_name, dic_new)


def turn_object_2_dict(dic, dic_new, parent, split):
    i=1
    for key in dic:
        i = i + 1
        value = dic[key]
        if isinstance(value, dict):
            new_parent=parent + split + key
            turn_object_2_dict(value, dic_new, new_parent, split)
            continue
        if parent != None and len(parent) != 0:
            dic_new[parent[1:len(parent)] + split + key] = value
        else:
            dic_new[key] = value

def turn_dict_2_object(dic, dicnew, parent, split):
    i=1
    for key in dic:
        i = i + 1
        value = dic[key]
        if isinstance(value, dict):
            new_parent=parent + split + key
            turn_dict_2_object(dic, dicnew, new_parent, split)
            continue
        if parent != None:
            dicnew[parent + split + key] = value
        else:
            dicnew[key] = value


def get_map_from_excel(obj):
    airport_df = load_workbook("all_activity.xlsx")
    sheets = airport_df.sheetnames

    sheet_first = sheets[0]

    ws2 = airport_df["Sheet2"]#读取excel中第二张表
    movement=[]
    name_mov=[]

    for i in range(1,ws2.max_column+1):
        for j in range(2,ws2.max_row+1):
            cellr=ws2.cell(j,i)
            movement.append(cellr.value)

    ln_movement=[]
    for j in range(len(movement)):
        ln_movement.append(math.log(movement[j]))###计算对数值


    mt=DataFrame(ln_movement,index=range(1,631),columns=["movement"])
    mt.describe()